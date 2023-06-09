import json
import sys

import openpyxl
from openpyxl.styles import Font, Alignment
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from datetime import datetime


def wait_for_page_to_load(driver):
    wait = WebDriverWait(driver, 30)
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")


def wait_for_element_presence(driver, xpath, timeout=30):
    try:
        element_present = EC.element_to_be_clickable((By.XPATH, xpath))
        WebDriverWait(driver, timeout).until(element_present)
    except TimeoutException:
        print(f"O elemento com xpath {xpath} não foi encontrado na página após {timeout} segundos de espera.")
        driver.quit()
        sys.exit()


def click_element(driver, xpath):
    wait_for_page_to_load(driver)
    wait_for_element_presence(driver, xpath)
    driver.find_element(By.XPATH, xpath).click()


def write_element(driver, xpath, msg):
    wait_for_page_to_load(driver)
    wait_for_element_presence(driver, xpath)
    elem = driver.find_element(By.XPATH, xpath)
    elem.clear()
    elem.send_keys(msg)


def read_attribute(driver, attribute, tag, text_element):
    wait_for_page_to_load(driver)
    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")
    element = soup.find(tag, string=text_element)
    att = element[attribute]
    return att


def find_first_available_day(driver, xpath_table, xpath_month, xpath_year):
    wait_for_page_to_load(driver)

    table = driver.find_element(By.XPATH, xpath_table)
    month = driver.find_element(By.XPATH, xpath_month).text
    year = driver.find_element(By.XPATH, xpath_year).text

    days = table.find_elements(By.TAG_NAME, 'td')
    first_available_day = None
    for day in days:
        if not day.get_attribute('class') or 'ui-state-disabled' not in day.get_attribute('class'):
            first_available_day = day.text
            break

    if first_available_day is None:
        return first_available_day
    else:
        return first_available_day + '/' + month + '/' + year


def save_date(date):
    # Formata data
    f_date = format_date(date)

    # Cria um dicionário principal
    data_dict = {}

    # Adiciona as datas ao dicionário principal
    data_dict["Instante_da_Coleta"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    data_dict["Data_Disponivel"] = f_date

    # Lê os dados existentes do arquivo JSON
    try:
        with open('data.json', 'r') as f:
            existing_data = json.load(f)
    except FileNotFoundError:
        existing_data = []

    # Adiciona o novo dicionário ao final da lista existente
    existing_data.append(data_dict)

    # Grava os dados atualizados no arquivo JSON
    with open('data.json', 'w') as f:
        json.dump(existing_data, f, indent=4)


def save_date_excel(date):
    if date == 'erro':
        f_date = 'A Página não carregou!'
    else:
        # Formata data
        f_date = format_date(date)

    # Abre a planilha existente
    try:
        wb = openpyxl.load_workbook(filename="data.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        # Se a planilha não existir, cria uma nova
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet["A1"] = "Instante da Coleta"
        sheet["B1"] = "Data Disponível"

    # Encontra a última linha preenchida
    last_row = sheet.max_row + 1

    # Adiciona as datas à planilha
    sheet.cell(row=last_row, column=1, value=datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    sheet.cell(row=last_row, column=2, value=f_date)

    # Define o estilo das células
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20

    title_font = Font(name='Arial', size=12, bold=True)
    title_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='center', vertical='center')

    # Aplica estilo apenas para as células do título
    for col in sheet.iter_cols(min_row=1, max_row=1, max_col=2):
        for cell in col:
            cell.font = title_font
            cell.alignment = title_alignment

    # Aplica estilo apenas para as células dos dados
    for row in sheet.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.alignment = data_alignment

    # Salva a planilha
    wb.save("data.xlsx")


def format_date(date):
    dt = datetime.strptime(date, '%d/%B/%Y')
    result = dt.strftime('%d/%m/%Y')
    return result
